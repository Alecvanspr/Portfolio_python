from venv import create
from xmlrpc.client import DateTime
import openpyxl
import pandas as pd
import re
from .models import Bestelling, Orderline

#TODOs
#de namen geniek maken naar de datum

#dit moet getest worden met behulp van intergratie in de webapp
class Path():
    def __init__(self, path = "NoName"):
         self.path  =  path
    def get(self):
        return self.path
    def set(self,nieuwPath):
        self.path = nieuwPath 

path = Path()
laatste_Regel = 24
___wb = openpyxl.load_workbook(filename='polls/excelBestanden/facturen/Template.xlsx')

def createFileName(id):
    bestelling = Bestelling.objects.get(id=id)
    datum = bestelling.lever_Datum
    user = bestelling.besteller.username
    return "polls/excelBestanden/facturen/"+user+"_"+str(datum) + ".xlsx"

def getTemplateWorkbook():
    return ___wb

def downloadFactuur(bestellingId):
    path.set(createFileName(bestellingId))
    if setDefaultValues(bestellingId):
        if fillOrderlines(bestellingId):
            print("Gelukt")
            return True
    return False

#met deze methode worden de normale gegevens ingevuld
def setDefaultValues(bestellingId):
    #try:
        bestelling = Bestelling.objects.get(id=bestellingId)
        workbook = getTemplateWorkbook()
        sheet = workbook.active
        #hier wordt de factuurdatum etc aangepast
        sheet['C21'] = bestelling.id
        sheet['F21'] = bestelling.factuur_Datum
        sheet['I21'] = bestelling.lever_Datum

        workbook.save(path.get())  
        print("bestellingsinformatie is ingevoerd")
        return True  
    #except:
     #   print("***** Bestellingsinformatie invoeren ging fout *****")
      #  return False

def fillOrderlines(id):
    try:
        regel = 24
        bestelling = Bestelling.objects.get(id=id)
        orderlines = Orderline.objects.filter(bestelling=bestelling)
        workbook = getTemplateWorkbook()
        sheet = workbook.active
        
        totaal = 0
        btw = 0

        for line in orderlines:
            r = str(regel)
            ProductTotaalPrijs =line.product.prijs*line.aantal
            sheet.merge_cells('C'+r+':G'+r)
            sheet['C'+r] = line.product.naam
            sheet["H"+r] = str(line.aantal)
            sheet["I"+r] = str(line.product.prijs)
            sheet["J"+r] = "9%"
            sheet["K"+r] = str(ProductTotaalPrijs)

            totaal+= ProductTotaalPrijs
            btw += 0
            regel+=1

        sheet.merge_cells('H'+str(regel)+":I"+str(regel))
        sheet.merge_cells('H'+str(regel+1)+":I"+str(regel+1))
        sheet.merge_cells('H'+str(regel+2)+":I"+str(regel+2))
        
        sheet['H'+str(regel)] = "Subtotaal"
        sheet['J'+str(regel)] = "=J"+str(regel+2)+"-J"+str(regel+1)

        
        sheet['H'+str(regel+1)] = "BTW"
        sheet['J'+str(regel+1)] = "=J"+str(regel+2)+"*0.09"
    
        
        sheet['H'+str(regel+2)] = "Totaal"
        sheet['J'+str(regel+2)] =totaal

        workbook.save(path.get())
        print("Orderlines zijn gemaakt")
        return True
    except:
        print("Het maken van de orderlines is gefaald")
        return False


#Deze methode is tijdelijk voor het testen enzo
def SlaOp(fileNaam):
    print("Wek")



